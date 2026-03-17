"""
Nexus Executor v3.0 — Multi-agent orchestrator with parallel execution.

Usage:
    python executor.py                         # Run one task and exit
    python executor.py --loop                  # Poll continuously (daemon mode)
    python executor.py --loop --workers 3      # Run 3 tasks in parallel
    python executor.py --loop --interval 30    # Poll every 30 seconds

Features (v3.0):
  - Parallel execution via ThreadPoolExecutor (--workers flag)
  - Agent-to-agent handoff (chain_next in task input_data)
  - Shared memory (swarm_memory table — store/recall across tasks)
  - Cost optimization (auto-route to haiku/sonnet/opus by task type)
  - Agent specialization (track success/fail per project+task_type)
  - Context library auto-loading from contexts/ directory
  - Approval gates (pending_approval status)
  - File I/O (xlsx, pdf, csv, text)
"""

import argparse
import json
import os
import platform
import subprocess
import sys
import threading
import time
import urllib.request
import urllib.error
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timezone

# ── Config ────────────────────────────────────────────────────────────

SUPABASE_URL = os.environ.get(
    "SUPABASE_URL", "https://ytvtaorgityczrdhhzqv.supabase.co"
)
SUPABASE_KEY = os.environ.get(
    "SUPABASE_KEY",
    "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6Inl0dnRhb3JnaXR5Y3pyZGhoenF2Iiwicm9sZSI6ImFub24iLCJpYXQiOjE3NzA5MzY4MTEsImV4cCI6MjA4NjUxMjgxMX0.A2uG-yVQ1HSV9-zlNDAztHHVw25g1cQ43180y3TfwGk",
)
NEXUS_URL = os.environ.get("NEXUS_URL", "http://localhost:3000")
DISCORD_WEBHOOK = os.environ.get("DISCORD_WEBHOOK_URL", "")
CLAUDE_CLI = os.environ.get("CLAUDE_CLI_PATH", "C:/nvm4w/nodejs/claude.cmd")
PROJECTS_ROOT = "C:/Users/Kruz/Desktop/Projects"
TASK_TIMEOUT = 600  # 10 minutes max per task
MAX_FILE_SIZE = 100 * 1024  # 100KB per file
OUTPUT_DIR = f"{PROJECTS_ROOT}/nexus/output"
CONTEXTS_DIR = f"{PROJECTS_ROOT}/nexus/contexts"

# Stable worker ID — survives restarts, unique per machine
HOSTNAME = platform.node().lower().replace(" ", "-")[:16]
WORKER_ID = f"executor-{HOSTNAME}"
WORKER_NAME = f"heavy-{HOSTNAME[:8]}"

# Project → directory mapping
PROJECT_DIRS = {
    "nexus": f"{PROJECTS_ROOT}/nexus",
    "MoneyPrinter": f"{PROJECTS_ROOT}/MoneyPrinter",
    "ai-finance-brief": f"{PROJECTS_ROOT}/ai-finance-brief",
    "ai-chess-coach": f"{PROJECTS_ROOT}/ai-chess-coach",
    "trade-journal": f"{PROJECTS_ROOT}/trade-journal",
    "buildkit-services": f"{PROJECTS_ROOT}/buildkit-services",
    "pc-bottleneck-analyzer": f"{PROJECTS_ROOT}/pc-bottleneck-analyzer",
    "outdoor-crm": f"{PROJECTS_ROOT}/outdoor-crm",
    "BarrelHouseCRM": f"{PROJECTS_ROOT}/BarrelHouseCRM",
    "email-finder-app": f"{PROJECTS_ROOT}/email-finder-app",
    "admin-dashboard": f"{PROJECTS_ROOT}/admin-dashboard",
    "n16-soccer": f"{PROJECTS_ROOT}/n16-soccer",
    "portfolio": f"{PROJECTS_ROOT}/portfolio",
    "mcp-servers": f"{PROJECTS_ROOT}/mcp-servers",
    "lead-tracker": f"{PROJECTS_ROOT}/lead-tracker",
    "wavefront": f"{PROJECTS_ROOT}/wavefront",
    "stock-breakout-alerts": f"{PROJECTS_ROOT}/stock-breakout-alerts",
    "municipal-crm": f"{PROJECTS_ROOT}/municipal-crm",
}

# Model routing: cost_tier → claude model flag
MODEL_ROUTING = {
    "haiku": "claude-haiku-4-5",
    "sonnet": "claude-sonnet-4-5",
    "light": "claude-haiku-4-5",
    # "cc_light", "heavy", None → default (no --model flag, uses Opus)
}

# Task types that are truly trivial — OK for Haiku
HAIKU_TASK_TYPES = {"eval", "check", "health", "status", "ping"}
# Task types that need real work — always Sonnet minimum
SONNET_TASK_TYPES = {"build", "scout", "builder", "implement", "fix", "refactor",
                     "review", "analyze", "plan", "research", "audit", "inspector", "miner"}
# Keywords for auto-classifying when task_type doesn't resolve
HAIKU_KEYWORDS = {"status", "ping", "health"}
SONNET_KEYWORDS = {"review", "analyze", "plan", "summarize", "compare", "audit", "research",
                   "build", "implement", "fix", "refactor", "scout"}
# Everything else → Opus (default)

# Thread safety
_fetch_lock = threading.Lock()
_stats_lock = threading.Lock()
_session_stats = {"completed": 0, "failed": 0, "xp": 0}


# ── Supabase helpers ──────────────────────────────────────────────────

def supabase_request(method, path, data=None):
    """Make a request to Supabase REST API."""
    url = f"{SUPABASE_URL}/rest/v1/{path}"
    headers = {
        "apikey": SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
        "Content-Type": "application/json",
        "Prefer": "return=representation",
    }

    body = json.dumps(data).encode() if data else None
    req = urllib.request.Request(url, data=body, headers=headers, method=method)

    try:
        with urllib.request.urlopen(req, timeout=10) as resp:
            return json.loads(resp.read().decode())
    except urllib.error.HTTPError as e:
        err_body = e.read().decode() if e.fp else str(e)
        print(f"  Supabase error {e.code}: {err_body}")
        return None
    except Exception as e:
        print(f"  Supabase request failed: {e}")
        return None


def update_task(task_id, updates):
    supabase_request("PATCH", f"swarm_tasks?id=eq.{task_id}", updates)


def log_task_event(task_id, event_type, title, project="", details=""):
    supabase_request("POST", "swarm_task_log", {
        "task_id": task_id,
        "event": event_type,
        "details": f"[{project}] {title}" + (f"\n{details}" if details else ""),
    })


def send_heartbeat(agent_name, project, status, step, steps_done=0, total=1, output=None):
    try:
        data = json.dumps({
            "agent_id": f"executor-{os.getpid()}",
            "agent_name": agent_name,
            "project": project,
            "status": status,
            "current_step": step,
            "steps_completed": steps_done,
            "total_steps": total,
            "output": output,
        }).encode()
        req = urllib.request.Request(
            f"{NEXUS_URL}/api/heartbeat", data=data,
            headers={"Content-Type": "application/json"}, method="POST",
        )
        urllib.request.urlopen(req, timeout=3)
    except Exception:
        pass


def notify_discord(message, color=0x06b6d4):
    if not DISCORD_WEBHOOK:
        return
    try:
        data = json.dumps({
            "username": "NEXUS Executor",
            "embeds": [{
                "description": message, "color": color,
                "footer": {"text": "NEXUS Executor v3"},
                "timestamp": datetime.now(timezone.utc).isoformat(),
            }],
        }).encode()
        req = urllib.request.Request(
            DISCORD_WEBHOOK, data=data,
            headers={"Content-Type": "application/json"}, method="POST",
        )
        urllib.request.urlopen(req, timeout=5)
    except Exception:
        pass


# ── Worker self-registration ─────────────────────────────────────────

def register_worker():
    now = datetime.now(timezone.utc).isoformat()
    worker_data = {
        "id": WORKER_ID, "worker_name": WORKER_NAME, "worker_type": "heavy",
        "tier": "executor", "status": "idle", "current_task_id": None,
        "last_heartbeat": now, "pid": os.getpid(),
        "tasks_completed": 0, "tasks_failed": 0,
        "total_cost_cents": 0, "total_tokens": 0, "xp": 0,
        "spawned_at": now, "died_at": None,
    }
    result = supabase_request("PATCH", f"swarm_workers?id=eq.{WORKER_ID}", {
        "status": "idle", "last_heartbeat": now, "pid": os.getpid(), "died_at": None,
    })
    if result is None or (isinstance(result, list) and len(result) == 0):
        supabase_request("POST", "swarm_workers", worker_data)
        print(f"  Registered as worker: {WORKER_ID}")
    else:
        print(f"  Reconnected as worker: {WORKER_ID}")


def update_worker(status="idle", task_id=None, project=None):
    with _stats_lock:
        stats = dict(_session_stats)
    supabase_request("PATCH", f"swarm_workers?id=eq.{WORKER_ID}", {
        "status": status, "current_task_id": task_id,
        "last_heartbeat": datetime.now(timezone.utc).isoformat(),
        "tasks_completed": stats["completed"], "tasks_failed": stats["failed"],
        "xp": stats["xp"],
    })


def deregister_worker():
    supabase_request("PATCH", f"swarm_workers?id=eq.{WORKER_ID}", {
        "status": "idle", "current_task_id": None,
        "last_heartbeat": datetime.now(timezone.utc).isoformat(),
    })
    print(f"  Worker {WORKER_ID} marked idle")


# ── Shared Memory (Feature 3) ───────────────────────────────────────

def memory_store(project, task_id, task_title, task_type, output_summary):
    """Store task output in shared memory for future tasks to recall."""
    supabase_request("POST", "swarm_memory", {
        "project": project,
        "task_id": task_id,
        "task_title": task_title,
        "task_type": task_type,
        "output_summary": output_summary[:3000],
    })


def memory_recall(project, limit=5):
    """Recall recent memory entries for a project."""
    result = supabase_request(
        "GET",
        f"swarm_memory?project=eq.{project}&order=created_at.desc&limit={limit}"
        f"&select=task_title,task_type,output_summary,created_at"
    )
    if not result:
        return ""

    parts = []
    for m in result:
        age = m.get("created_at", "")[:10]
        parts.append(f"- [{age}] {m.get('task_title', '?')}: {m.get('output_summary', '')[:200]}")

    if parts:
        return "Recent project memory:\n" + "\n".join(parts) + "\n\n"
    return ""


# ── Cost Optimization / Model Routing (Feature 6) ───────────────────

def select_model(task):
    """Choose the right Claude model based on task complexity.

    Routing priority:
      1. task_type in HAIKU_TASK_TYPES  → Haiku  (trivial ops only)
      2. task_type in SONNET_TASK_TYPES → Sonnet (real work)
      3. Keyword scan of title/description as fallback
      4. Default → Sonnet (safe default; Opus only when explicitly requested)

    cost_tier field is IGNORED for routing — it was sending real work to
    Haiku.  We keep the field for billing/reporting only.
    """
    task_type = (task.get("task_type") or "").lower().strip()

    # 1. Route by task_type first (most reliable signal)
    if task_type in HAIKU_TASK_TYPES:
        return MODEL_ROUTING["haiku"]
    if task_type in SONNET_TASK_TYPES:
        return MODEL_ROUTING["sonnet"]

    # 2. Keyword fallback for unknown task_types
    text = (task.get("title", "") + " " + (task.get("description") or "")).lower()
    words = set(text.split())

    if words & HAIKU_KEYWORDS and not words & SONNET_KEYWORDS:
        return MODEL_ROUTING["haiku"]

    if words & SONNET_KEYWORDS:
        return MODEL_ROUTING["sonnet"]

    # 3. Default to Sonnet — safe for real work, not wasteful
    return MODEL_ROUTING["sonnet"]


# ── Agent Specialization (Features 4 & 5) ───────────────────────────

def load_specialization(project, task_type):
    """Load best practices and patterns for this project+task_type."""
    result = supabase_request(
        "GET",
        f"agent_specializations?project=eq.{project}&task_type=eq.{task_type}&limit=1"
    )
    if result and len(result) > 0:
        spec = result[0]
        parts = []
        if spec.get("best_practices"):
            parts.append(f"Best practices for {project}/{task_type}:\n{spec['best_practices']}")
        if spec.get("common_errors"):
            parts.append(f"Common errors to avoid:\n{spec['common_errors']}")
        success_rate = 0
        total = (spec.get("success_count") or 0) + (spec.get("fail_count") or 0)
        if total > 0:
            success_rate = round((spec.get("success_count", 0) / total) * 100)
            parts.append(f"Historical success rate: {success_rate}% ({total} tasks)")
        if parts:
            return "\n".join(parts) + "\n\n"
    return ""


def track_specialization(project, task_type, success, duration):
    """Update specialization stats after task completion."""
    # Upsert — try PATCH first, then POST
    field = "success_count" if success else "fail_count"

    existing = supabase_request(
        "GET",
        f"agent_specializations?project=eq.{project}&task_type=eq.{task_type}&limit=1"
    )

    now = datetime.now(timezone.utc).isoformat()

    if existing and len(existing) > 0:
        row = existing[0]
        new_count = (row.get(field) or 0) + 1
        total = (row.get("success_count") or 0) + (row.get("fail_count") or 0) + 1
        old_avg = row.get("avg_duration_seconds") or 0
        new_avg = round(((old_avg * (total - 1)) + duration) / total, 1)

        supabase_request("PATCH", f"agent_specializations?id=eq.{row['id']}", {
            field: new_count,
            "avg_duration_seconds": new_avg,
            "last_updated": now,
        })
    else:
        supabase_request("POST", "agent_specializations", {
            "project": project,
            "task_type": task_type,
            "success_count": 1 if success else 0,
            "fail_count": 0 if success else 1,
            "avg_duration_seconds": round(duration, 1),
            "last_updated": now,
        })


# ── Agent-to-Agent Handoff (Feature 1) ──────────────────────────────

def handle_handoff(task, output):
    """After a task completes, spawn the next task in the chain if defined."""
    input_data = task.get("input_data") or {}
    if isinstance(input_data, str):
        try:
            input_data = json.loads(input_data)
        except Exception:
            return

    chain_next = input_data.get("chain_next")
    if not chain_next or not isinstance(chain_next, list):
        return

    task_id = task["id"]
    project = task.get("project", "general")
    now = datetime.now(timezone.utc).isoformat()

    for i, next_step in enumerate(chain_next):
        next_id = f"{task_id[:8]}-chain-{i}"
        goal = next_step.get("description", next_step.get("goal", ""))

        # Inject parent output into the next step
        if next_step.get("use_parent_output", True):
            goal = f"Previous step output:\n\n{output[:3000]}\n\n---\n\n{goal}"

        new_task = {
            "id": next_id,
            "title": next_step.get("title", f"[Handoff] {next_step.get('task_type', 'review')}"),
            "description": goal,
            "project": next_step.get("project", project),
            "priority": next_step.get("priority", 30),
            "status": "queued",
            "task_type": next_step.get("task_type", "inspector"),
            "cost_tier": next_step.get("cost_tier", "cc_light"),
            "created_at": now,
            "updated_at": now,
            "parent_task_id": task_id,
            "retry_count": 0,
            "max_retries": 2,
        }

        supabase_request("POST", "swarm_tasks", new_task)
        log_task_event(next_id, "handoff", f"Auto-spawned from {task_id[:8]}", project)
        print(f"  HANDOFF: Spawned {next_step.get('task_type', 'task')} → {next_id[:12]}")

    notify_discord(
        f"**Agent Handoff:** {task.get('title', '')[:50]}\n"
        f"Spawned {len(chain_next)} follow-up task(s)",
        0x8b5cf6,
    )


def unblock_dependents(task_id):
    """Move blocked tasks that depend on this completed task to queued."""
    # Find tasks that are blocked and depend on this task
    blocked = supabase_request(
        "GET",
        f"swarm_tasks?status=eq.blocked&select=id,depends_on"
    )
    if not blocked:
        return

    for t in blocked:
        deps = t.get("depends_on") or []
        if task_id in deps:
            # Check if ALL dependencies are completed
            all_done = True
            for dep_id in deps:
                if dep_id == task_id:
                    continue
                dep_task = supabase_request("GET", f"swarm_tasks?id=eq.{dep_id}&select=status&limit=1")
                if dep_task and len(dep_task) > 0 and dep_task[0].get("status") != "completed":
                    all_done = False
                    break

            if all_done:
                update_task(t["id"], {
                    "status": "queued",
                    "updated_at": datetime.now(timezone.utc).isoformat(),
                })
                print(f"  UNBLOCKED: {t['id'][:8]} (dependency {task_id[:8]} completed)")


# ── File I/O ─────────────────────────────────────────────────────────

def read_xlsx(file_path):
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        parts = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            for row in ws.iter_rows(values_only=True):
                rows.append(",".join(str(c) if c is not None else "" for c in row))
            if rows:
                parts.append(f"=== Sheet: {sheet_name} ({len(rows)} rows) ===\n" + "\n".join(rows[:500]))
        wb.close()
        return "\n\n".join(parts)
    except ImportError:
        return "[openpyxl not installed]"
    except Exception as e:
        return f"[Error reading xlsx: {e}]"


def read_pdf(file_path):
    try:
        from PyPDF2 import PdfReader
        reader = PdfReader(file_path)
        pages = []
        for i, page in enumerate(reader.pages[:50]):
            text = page.extract_text() or ""
            if text.strip():
                pages.append(f"--- Page {i+1} ---\n{text}")
        return "\n\n".join(pages) if pages else "[PDF has no extractable text]"
    except ImportError:
        return "[PyPDF2 not installed]"
    except Exception as e:
        return f"[Error reading pdf: {e}]"


def save_output(task_id, project, content):
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    filename = f"{project}_{task_id[:8]}_{timestamp}.md"
    filepath = os.path.join(OUTPUT_DIR, filename)
    try:
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(content)
        return filepath
    except Exception:
        return None


def load_context_files(project, task_type=None):
    if not os.path.isdir(CONTEXTS_DIR):
        return ""
    context_parts = []
    for filename in sorted(os.listdir(CONTEXTS_DIR)):
        if not filename.endswith('.md'):
            continue
        name_lower = filename.lower()
        is_global = not any(name_lower.startswith(p) for p in [
            "nexus-", "moneyprinter-", "deere-", "buildkit-", "finance-"])
        is_project = name_lower.startswith(f"{project.lower()}-") or name_lower.startswith("deere-")
        is_finance = name_lower.startswith("finance-") and task_type in ("scout", "inspector", "miner")
        if is_global or is_project or is_finance:
            filepath = os.path.join(CONTEXTS_DIR, filename)
            try:
                with open(filepath, 'r', encoding='utf-8') as f:
                    content = f.read()
                if len(content) > 10000:
                    content = content[:10000] + "\n[...truncated...]"
                context_parts.append(f"--- CONTEXT: {filename} ---\n{content}\n--- END CONTEXT ---")
            except Exception:
                pass
    return "\n\n".join(context_parts) + "\n\n" if context_parts else ""


def read_input_files(input_data):
    files = input_data.get("files", [])
    if not files:
        return ""
    context_parts = []
    for file_path in files:
        try:
            if not os.path.isfile(file_path):
                context_parts.append(f"[File not found: {file_path}]")
                continue
            size = os.path.getsize(file_path)
            if size > MAX_FILE_SIZE:
                context_parts.append(f"[File too large ({size} bytes): {file_path}]")
                continue
            ext = os.path.splitext(file_path)[1].lower()
            if ext in ('.xlsx', '.xls'):
                context_parts.append(f"--- SPREADSHEET: {file_path} ---\n{read_xlsx(file_path)}\n--- END ---")
            elif ext == '.pdf':
                context_parts.append(f"--- PDF: {file_path} ---\n{read_pdf(file_path)}\n--- END ---")
            elif ext in ('.txt', '.csv', '.json', '.md', '.py', '.ts', '.tsx', '.js', '.sql', '.yaml', '.yml', '.toml', '.env', '.log', '.html', '.xml'):
                with open(file_path, 'r', encoding='utf-8', errors='replace') as f:
                    content = f.read()
                context_parts.append(f"--- FILE: {file_path} ---\n{content}\n--- END ---")
            else:
                context_parts.append(f"[Unsupported file type: {file_path}]")
        except Exception as e:
            context_parts.append(f"[Error reading {file_path}: {e}]")
    return "\n\n".join(context_parts) + "\n\n" if context_parts else ""


# ── Task Fetching (thread-safe) ──────────────────────────────────────

def fetch_next_task():
    """Get the highest-priority queued or approved task (thread-safe).

    Uses optimistic locking: PATCH includes status=eq.queued guard so only
    one process can claim a task even without an RPC.
    """
    with _fetch_lock:
        # Approved tasks first
        result = supabase_request("GET", "swarm_tasks?status=eq.approved&order=priority.asc&limit=1")
        if result and len(result) > 0:
            task = result[0]
            # Optimistic lock: only claim if still approved (prevents double-execution)
            claimed = supabase_request("PATCH",
                f"swarm_tasks?id=eq.{task['id']}&status=eq.approved", {
                    "status": "running",
                    "updated_at": datetime.now(timezone.utc).isoformat(),
                    "assigned_worker_id": WORKER_ID,
                })
            if claimed and len(claimed) > 0:
                return task
            # Another worker claimed it, fall through

        # Queued tasks
        result = supabase_request("GET", "swarm_tasks?status=eq.queued&order=priority.asc&limit=1")
        if result and len(result) > 0:
            task = result[0]
            # Optimistic lock: only claim if still queued
            claimed = supabase_request("PATCH",
                f"swarm_tasks?id=eq.{task['id']}&status=eq.queued", {
                    "status": "running",
                    "updated_at": datetime.now(timezone.utc).isoformat(),
                    "assigned_worker_id": WORKER_ID,
                })
            if claimed and len(claimed) > 0:
                return task
            # Another worker claimed it

        return None


# ── Approval Gates ───────────────────────────────────────────────────

def check_approval_gate(task):
    input_data = task.get("input_data") or {}
    if isinstance(input_data, str):
        try:
            input_data = json.loads(input_data)
        except Exception:
            input_data = {}

    needs_approval = input_data.get("wait_for_approval", False)
    title = task.get("title", "")
    if "[approval]" in title.lower():
        needs_approval = True

    if not needs_approval:
        return True
    if task.get("status") == "approved":
        return True

    task_id = task["id"]
    project = task.get("project", "general")
    update_task(task_id, {"status": "pending_approval", "updated_at": datetime.now(timezone.utc).isoformat()})
    log_task_event(task_id, "approval_pending", f"Awaiting approval: {title}", project)
    notify_discord(
        f"**Approval Required**\nTask: {title}\nProject: {project}\n\n"
        f"Approve: `POST /api/tasks/approve` with `{{\"task_id\": \"{task_id}\"}}`",
        0xe8a019,
    )
    print(f"  APPROVAL REQUIRED: {title}")
    return False


# ── Task Execution ───────────────────────────────────────────────────

def execute_task(task):
    """Execute a single task via Claude Code CLI. Thread-safe."""
    task_id = task["id"]
    title = task.get("title", "Untitled")
    project = task.get("project", "general")
    task_type = task.get("task_type", "builder")
    prompt = task.get("description") or task.get("title", "")

    # Check approval gate
    if not check_approval_gate(task):
        return True  # Not a failure, just waiting

    # === Intelligence layers (Month 3) ===

    # 1. Load agent specialization / best practices
    spec_context = load_specialization(project, task_type)
    if spec_context:
        prompt = f"Agent specialization notes:\n{spec_context}\n{prompt}"

    # 2. Recall shared memory for this project
    memory_context = memory_recall(project, limit=3)
    if memory_context:
        prompt = f"{memory_context}{prompt}"

    # 3. Auto-load context files
    context = load_context_files(project, task_type)
    if context:
        prompt = f"Reference context:\n\n{context}\nTask:\n{prompt}"

    # 4. Read input files
    input_data = task.get("input_data") or {}
    if isinstance(input_data, str):
        try:
            input_data = json.loads(input_data)
        except Exception:
            input_data = {}
    file_context = read_input_files(input_data)
    if file_context:
        prompt = f"Input files:\n\n{file_context}\n{prompt}"

    # Ensure input_data.prompt is populated so swarm workers can read it
    # (workers like light_worker, heavy_worker expect input_data.prompt)
    input_data["prompt"] = prompt
    task["input_data"] = input_data

    # Resolve working directory
    cwd = PROJECT_DIRS.get(project, PROJECTS_ROOT)
    if not os.path.isdir(cwd):
        cwd = PROJECTS_ROOT

    thread_name = threading.current_thread().name
    print(f"\n{'='*60}")
    print(f"  [{thread_name}] EXECUTING: {title}")
    print(f"  Project: {project} | Type: {task_type} | ID: {task_id[:8]}")
    print(f"{'='*60}\n")

    # Task is already marked running by fetch_next_task
    now = datetime.now(timezone.utc).isoformat()
    update_task(task_id, {"started_at": now})
    log_task_event(task_id, "started", f"Executor started: {title}", project)
    send_heartbeat(f"Executor: {title[:40]}", project, "running", "Starting Claude Code...", 0, 3)
    update_worker("busy", task_id, project)

    # 5. Model routing (cost optimization)
    model = select_model(task)
    print(f"  Prompt length: {len(prompt)} chars")

    # Write prompt to temp file to avoid Windows .cmd argument mangling
    import tempfile
    prompt_file = os.path.join(tempfile.gettempdir(), f"nexus-prompt-{task_id[:8]}.txt")
    with open(prompt_file, "w", encoding="utf-8") as pf:
        pf.write(prompt)

    # Build shell command that pipes the prompt file to claude CLI
    model_flag = f' --model {model}' if model else ""
    shell_cmd = f'type "{prompt_file}" | "{CLAUDE_CLI}" --output-format text -p - --dangerously-skip-permissions{model_flag}'
    if model:
        print(f"  Model: {model}")

    try:
        start = time.time()
        # Hide cmd.exe windows on Windows so they don't spam the desktop
        startupinfo = None
        if os.name == "nt":
            startupinfo = subprocess.STARTUPINFO()
            startupinfo.dwFlags |= subprocess.STARTF_USESHOWWINDOW
            startupinfo.wShowWindow = subprocess.SW_HIDE
        result = subprocess.run(
            shell_cmd, cwd=cwd, capture_output=True, text=True,
            timeout=TASK_TIMEOUT, shell=True, encoding="utf-8", errors="replace",
            startupinfo=startupinfo,
        )
        duration = round(time.time() - start, 1)

        stdout = result.stdout or ""
        stderr = result.stderr or ""

        if len(stdout) > 15000:
            stdout = stdout[:5000] + f"\n\n[...truncated {len(stdout) - 10000} chars...]\n\n" + stdout[-5000:]

        if result.returncode == 0:
            # ── Clarification detection ──────────────────────────
            # If the model asked questions instead of doing work,
            # treat it as a failure so the task gets retried or
            # escalated instead of silently "completing".
            CLARIFICATION_PHRASES = [
                "i need more information",
                "i need more details",
                "i need more context",
                "could you clarify",
                "could you provide",
                "could you specify",
                "can you clarify",
                "can you provide more",
                "need specific direction",
                "need clarification",
                "i'm ready to work but",
                "i'm ready to help but",
                "please provide",
                "please clarify",
                "before i proceed",
                "before i can",
                "what would you like me",
                "which approach would you prefer",
            ]
            stdout_lower = stdout.lower()
            asked_clarification = any(phrase in stdout_lower for phrase in CLARIFICATION_PHRASES)

            if asked_clarification:
                print(f"\n  [{thread_name}] REJECTED (clarification instead of execution) in {duration}s")

                with _stats_lock:
                    _session_stats["failed"] += 1
                    _session_stats["xp"] += 2

                error_reason = "Model requested clarification instead of executing"
                update_task(task_id, {
                    "status": "failed",
                    "updated_at": datetime.now(timezone.utc).isoformat(),
                    "error_message": error_reason,
                    "output_data": json.dumps({
                        "response": stdout, "duration_seconds": duration,
                        "exit_code": 0, "model": model or "default",
                        "rejection_reason": "clarification_instead_of_execution",
                    }),
                })
                log_task_event(task_id, "rejected", f"Clarification instead of work: {title}", project, error_reason)
                update_worker("idle")
                notify_discord(
                    f"**Task Rejected (clarification):** {title}\n"
                    f"Project: {project} | {duration}s | Model: {model or 'default'}",
                    0xe8a019,
                )
                track_specialization(project, task_type, False, duration)
                return False
            # ── End clarification detection ──────────────────────

            print(f"\n  [{thread_name}] COMPLETED in {duration}s")

            with _stats_lock:
                _session_stats["completed"] += 1
                _session_stats["xp"] += 10

            output_path = save_output(task_id, project, stdout) if len(stdout) > 100 else None

            update_task(task_id, {
                "status": "completed",
                "completed_at": datetime.now(timezone.utc).isoformat(),
                "updated_at": datetime.now(timezone.utc).isoformat(),
                "output_data": json.dumps({
                    "response": stdout, "duration_seconds": duration,
                    "exit_code": 0, "output_file": output_path,
                    "model": model or "default",
                }),
            })
            log_task_event(task_id, "completed", f"Completed in {duration}s: {title}", project)
            send_heartbeat(f"Executor: {title[:40]}", project, "completed", "Done!", 3, 3, stdout[:500])
            update_worker("idle")
            notify_discord(f"**Mission Complete:** {title}\nProject: {project} | {duration}s | Model: {model or 'opus'}", 0x10b981)

            # === Post-completion intelligence ===
            # Store in shared memory
            summary = stdout[:500] if len(stdout) > 500 else stdout
            memory_store(project, task_id, title, task_type, summary)

            # Track specialization
            track_specialization(project, task_type, True, duration)

            # Agent-to-agent handoff
            handle_handoff(task, stdout)

            # Unblock dependent tasks
            unblock_dependents(task_id)

            return True
        else:
            error_msg = stderr[:500] or f"Exit code {result.returncode}"
            print(f"\n  [{thread_name}] FAILED (exit {result.returncode}) in {duration}s")

            with _stats_lock:
                _session_stats["failed"] += 1
                _session_stats["xp"] += 2

            update_task(task_id, {
                "status": "failed",
                "updated_at": datetime.now(timezone.utc).isoformat(),
                "error_message": error_msg,
                "output_data": json.dumps({
                    "stdout": stdout[:5000], "stderr": stderr[:5000],
                    "duration_seconds": duration, "exit_code": result.returncode,
                }),
            })
            log_task_event(task_id, "failed", f"Failed: {title}", project, error_msg)
            update_worker("idle")
            notify_discord(f"**Mission Failed:** {title}\nProject: {project}\nError: {error_msg[:200]}", 0xef4444)

            # Track failure for specialization
            track_specialization(project, task_type, False, duration)

            return False

    except subprocess.TimeoutExpired as te:
        duration = round(time.time() - start, 1)
        print(f"\n  [{thread_name}] TIMED OUT after {TASK_TIMEOUT}s")
        # Kill orphan claude processes spawned by shell=True
        # subprocess.run only kills cmd.exe, not the grandchild claude process
        try:
            if os.name == "nt":
                # Kill any claude processes with our prompt file
                os.system(f'taskkill /F /FI "WINDOWTITLE eq nexus-prompt-{task_id[:8]}*" >nul 2>&1')
        except Exception:
            pass
        with _stats_lock:
            _session_stats["failed"] += 1
        update_task(task_id, {
            "status": "failed", "updated_at": datetime.now(timezone.utc).isoformat(),
            "error_message": f"Timed out after {TASK_TIMEOUT}s",
        })
        log_task_event(task_id, "timeout", f"Timed out: {title}", project)
        update_worker("idle")
        notify_discord(f"**Mission Timeout:** {title}\nProject: {project}", 0xe8a019)
        track_specialization(project, task_type, False, TASK_TIMEOUT)
        return False

    except Exception as e:
        print(f"\n  [{thread_name}] ERROR: {e}")
        with _stats_lock:
            _session_stats["failed"] += 1
        update_task(task_id, {
            "status": "failed", "updated_at": datetime.now(timezone.utc).isoformat(),
            "error_message": str(e)[:500],
        })
        log_task_event(task_id, "error", f"Error: {title}", project, str(e))
        update_worker("idle")
        return False

    finally:
        # Clean up temp prompt file
        try:
            if prompt_file and os.path.exists(prompt_file):
                os.remove(prompt_file)
        except Exception:
            pass


# ── Main ──────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Nexus Executor v3 — multi-agent orchestrator")
    parser.add_argument("--loop", action="store_true", help="Poll continuously (daemon mode)")
    parser.add_argument("--interval", type=int, default=15, help="Poll interval in seconds (default: 15)")
    parser.add_argument("--workers", type=int, default=3, help="Max parallel tasks (default: 3, max: 5)")
    args = parser.parse_args()
    args.workers = min(args.workers, 5)  # Cap at 5

    print(f"""
    ========================================
        NEXUS EXECUTOR v3.0
        Multi-agent orchestrator.
        Parallel | Handoff | Memory | Routing
    ========================================
    """)
    print(f"  Claude CLI:  {CLAUDE_CLI}")
    print(f"  Nexus URL:   {NEXUS_URL}")
    print(f"  Worker ID:   {WORKER_ID}")
    print(f"  Workers:     {args.workers} parallel")
    print(f"  Mode:        {'Daemon (loop)' if args.loop else 'Single run'}")
    print()

    if args.loop:
        register_worker()
        notify_discord(
            f"**Executor v3 Online:** `{WORKER_ID}`\n"
            f"Parallel workers: {args.workers} | Polling every {args.interval}s\n"
            f"Features: handoff, memory, model routing, specialization",
            0x06b6d4,
        )

        print(f"  Polling every {args.interval}s for queued tasks...")
        print(f"  Max {args.workers} tasks in parallel")
        print("  Press Ctrl+C to stop.\n")

        heartbeat_counter = 0
        pool = ThreadPoolExecutor(max_workers=args.workers, thread_name_prefix="agent")
        active_futures = set()

        try:
            while True:
                # Clean up completed futures
                done = {f for f in active_futures if f.done()}
                for f in done:
                    try:
                        f.result()  # Raise any exceptions
                    except Exception as e:
                        print(f"  Thread error: {e}")
                active_futures -= done

                # Submit new tasks if we have capacity
                while len(active_futures) < args.workers:
                    task = fetch_next_task()
                    if task:
                        future = pool.submit(execute_task, task)
                        active_futures.add(future)
                    else:
                        break  # No more tasks available

                if not active_futures:
                    sys.stdout.write(".")
                    sys.stdout.flush()

                # Heartbeat
                heartbeat_counter += 1
                if heartbeat_counter >= 4:
                    update_worker("idle" if not active_futures else "busy")
                    heartbeat_counter = 0

                time.sleep(args.interval)

        except KeyboardInterrupt:
            print("\n\n  Shutting down gracefully...")
            pool.shutdown(wait=True, cancel_futures=True)
            deregister_worker()
            notify_discord(f"**Executor Offline:** `{WORKER_ID}` — clean shutdown", 0xe8a019)
            print("  Executor stopped.")
    else:
        task = fetch_next_task()
        if task:
            success = execute_task(task)
            sys.exit(0 if success else 1)
        else:
            print("  No queued tasks found.")
            sys.exit(0)


if __name__ == "__main__":
    main()
